from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
from .models import Student,AdditionalInfo 
from .serializers import StudentSerializer ,AdditionalInfoSerializer

class Student(APIView):
    def post(self, request):
        try:
            data = request.FILES['file']
            workbook = openpyxl.load_workbook(data)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, age, roll, mobile, email, city, state = row
                student_data = {
                    'name': name,
                    'age': age,
                    'roll': roll,
                    'mobile': mobile,
                    'email': email,
                    'city': city,
                    'state': state
                }

                serializer = StudentSerializer(data=student_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "Data uploaded successfully"}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"message": f"Error processing file: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    
class AdditionalInfo(APIView):
    def post(self,request):
        serializer = AdditionalInfo(data=request.data)
        serializer2 = AdditionalInfo(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            if serializer2.is_valid(raise_exception=True):
                serializer2.save()
                return Response("inserted",status=status.HTTP_201_CREATED)
            return Response("inserted",status=status.HTTP_201_CREATED)
        return Response("no")
        

# class AdditionalInfo(APIView):
#     def post(self, request):
#         if 'file' in request.FILES:
#             file = request.FILES['file']
#             workbook = openpyxl.load_workbook(file)
#             sheet = workbook.active

#             for row in sheet.iter_rows(min_row=2, values_only=True):
#                 student_data = {
#                     'father_name': row[0],
#                     'father_mobile': row[1],
#                     'father_age': row[2],
#                     'father_email': row[3],
#                     # Other student fields...
#                 }

#                 student_serializer = AdditionalInfoSerializer(data=student_data)
#                 if student_serializer.is_valid():
#                     student = student_serializer.save()

#                     additional_info_data = {
#                         'student': student.id,
#                         'info_field': row[7],
                       
#                     }

#                     additional_info_serializer = AdditionalInfoSerializer(data=additional_info_data)
#                     if additional_info_serializer.is_valid():
#                         additional_info_serializer.save()
#                     else:
#                         return Response(additional_info_serializer.errors, status=400)
#                 else:
#                     return Response(student_serializer.errors, status=400)

#             return Response({"message": "Data uploaded successfully"}, status=201)
#         else:
#             return Response({"message": "No file found in the request"}, status=400)